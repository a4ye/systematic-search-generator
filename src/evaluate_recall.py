"""Evaluate recall of a .ris file against an included studies .xlsx file."""

import argparse
import re
import sys
from pathlib import Path

import openpyxl
from rich.console import Console
from rich.table import Table


def _normalize_doi(doi: str) -> str:
    """Normalize a DOI for consistent matching."""
    doi = doi.strip().rstrip(".")
    doi = re.sub(r"^https?://doi\.org/", "", doi, flags=re.IGNORECASE)
    prefix_end = doi.find("/")
    if prefix_end > 0:
        prefix = doi[:prefix_end]
        suffix = doi[prefix_end:]
        suffix = re.sub(r"/{2,}", "/", suffix)
        doi = prefix + suffix
    return doi.lower()


def parse_ris(path: Path) -> tuple[set[str], set[str], dict[str, str]]:
    """Parse a .ris file and return (pmids, dois, doi_to_pmid)."""
    pmids: set[str] = set()
    dois: set[str] = set()
    doi_to_pmid: dict[str, str] = {}

    current_pmid: str | None = None
    current_dois: list[str] = []

    for line in path.read_text().splitlines():
        line = line.strip()
        if line.startswith("ID  - "):
            current_pmid = line[6:].strip()
            pmids.add(current_pmid)
        elif line.startswith("DO  - "):
            doi = _normalize_doi(line[6:].strip())
            dois.add(doi)
            current_dois.append(doi)
        elif line.startswith("ER  -"):
            if current_pmid:
                for doi in current_dois:
                    doi_to_pmid[doi] = current_pmid
            current_pmid = None
            current_dois = []

    return pmids, dois, doi_to_pmid


def parse_included_studies(path: Path) -> list[dict]:
    """Parse an included studies .xlsx file.

    Returns a list of dicts with keys: pmid, doi, title, row.
    """
    wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
    ws = wb.active

    headers = [cell.value for cell in ws[1]]
    headers_lower = [str(h).lower().strip() if h else "" for h in headers]

    doi_col = None
    pmid_col = None
    title_col = None
    for i, h in enumerate(headers_lower):
        if h == "doi":
            doi_col = i
        elif h in ("pubmed id", "pmid", "pubmed_id"):
            pmid_col = i
        elif h == "title":
            title_col = i

    studies: list[dict] = []
    for row_num, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
        doi_raw = row[doi_col] if doi_col is not None and doi_col < len(row) else None
        pmid_raw = row[pmid_col] if pmid_col is not None and pmid_col < len(row) else None
        title_raw = row[title_col] if title_col is not None and title_col < len(row) else None

        doi = _normalize_doi(str(doi_raw)) if doi_raw else None
        pmid = str(int(pmid_raw)) if pmid_raw is not None and pmid_raw != "" else None
        title = str(title_raw).strip() if title_raw else None

        if not doi and not pmid and not title:
            continue

        studies.append({"pmid": pmid, "doi": doi, "title": title, "row": row_num})

    wb.close()
    return studies


def evaluate(
    ris_pmids: set[str],
    ris_dois: set[str],
    ris_doi_to_pmid: dict[str, str],
    studies: list[dict],
) -> dict:
    """Evaluate recall of .ris results against included studies.

    A study is considered PubMed-indexed if:
    - It has a PMID in the spreadsheet, OR
    - Its DOI maps to a PMID in the .ris file (meaning PubMed has it indexed)

    Returns a dict with metrics and per-study match results.
    """
    found: list[dict] = []
    missed: list[dict] = []

    for study in studies:
        matched = False
        match_type = None

        if study["pmid"] and study["pmid"] in ris_pmids:
            matched = True
            match_type = "pmid"
        elif study["doi"] and study["doi"] in ris_dois:
            matched = True
            match_type = "doi"

        # Determine if this study is PubMed-indexed:
        # either it has a PMID in the spreadsheet, or its DOI resolves to a
        # PMID in the .ris (which means PubMed indexes it under that PMID)
        has_pmid = bool(study["pmid"])
        if not has_pmid and study["doi"] and study["doi"] in ris_doi_to_pmid:
            has_pmid = True

        if matched:
            found.append({**study, "match_type": match_type, "pubmed_indexed": has_pmid})
        else:
            missed.append({**study, "pubmed_indexed": has_pmid})

    total = len(studies)
    all_with_flag = found + missed
    pubmed_indexed = [s for s in all_with_flag if s["pubmed_indexed"]]
    pubmed_found = [s for s in found if s["pubmed_indexed"]]

    found_count = len(found)
    recall_overall = found_count / total if total > 0 else 0
    recall_pubmed = len(pubmed_found) / len(pubmed_indexed) if pubmed_indexed else 0
    precision = found_count / len(ris_pmids) if ris_pmids else 0
    nnr = len(ris_pmids) / found_count if found_count > 0 else float("inf")

    return {
        "total": total,
        "found": found_count,
        "missed_count": len(missed),
        "pubmed_indexed": len(pubmed_indexed),
        "pubmed_found": len(pubmed_found),
        "recall_overall": recall_overall,
        "recall_pubmed": recall_pubmed,
        "precision": precision,
        "nnr": nnr,
        "ris_pmid_count": len(ris_pmids),
        "ris_doi_count": len(ris_dois),
        "found_studies": found,
        "missed_studies": missed,
    }


def main():
    parser = argparse.ArgumentParser(
        description="Evaluate recall of a .ris file against included studies."
    )
    parser.add_argument("ris_file", type=Path, help="Path to .ris file")
    parser.add_argument("included_file", type=Path, help="Path to included studies .xlsx file")
    args = parser.parse_args()

    console = Console()

    if not args.ris_file.exists():
        console.print(f"[red]RIS file not found: {args.ris_file}[/red]")
        sys.exit(1)
    if not args.included_file.exists():
        console.print(f"[red]Included studies file not found: {args.included_file}[/red]")
        sys.exit(1)

    ris_pmids, ris_dois, ris_doi_to_pmid = parse_ris(args.ris_file)
    studies = parse_included_studies(args.included_file)
    result = evaluate(ris_pmids, ris_dois, ris_doi_to_pmid, studies)

    # Metrics table
    metrics_table = Table(title="Recall Evaluation")
    metrics_table.add_column("Metric", style="bold")
    metrics_table.add_column("Value", justify="right")

    metrics_table.add_row("Included studies", str(result["total"]))
    metrics_table.add_row("  PubMed-indexed", str(result["pubmed_indexed"]))
    metrics_table.add_row("  No PMID", str(result["total"] - result["pubmed_indexed"]))
    metrics_table.add_row("RIS PMIDs", f"{result['ris_pmid_count']:,}")
    metrics_table.add_row("RIS DOIs", f"{result['ris_doi_count']:,}")
    metrics_table.add_row("Found", f"{result['found']} / {result['total']}")
    metrics_table.add_row("Recall (overall)", f"{result['recall_overall']:.1%}")
    metrics_table.add_row("Recall (PubMed)", f"{result['recall_pubmed']:.1%}")
    metrics_table.add_row("Precision", f"{result['precision']:.2%}")
    metrics_table.add_row("NNR", f"{result['nnr']:.1f}")

    console.print(metrics_table)

    # Missed studies
    missed = result["missed_studies"]
    if missed:
        console.print()
        missed_table = Table(title=f"Missed Studies ({len(missed)})")
        missed_table.add_column("#", justify="right")
        missed_table.add_column("PMID")
        missed_table.add_column("DOI")
        missed_table.add_column("Title", max_width=60)

        for i, study in enumerate(missed, 1):
            missed_table.add_row(
                str(i),
                study["pmid"] or "-",
                study["doi"] or "-",
                study["title"] or "-",
            )

        console.print(missed_table)
    else:
        console.print("\n[green]All included studies found![/green]")


if __name__ == "__main__":
    main()
